import tkinter as tk
from tkinter import ttk,messagebox
import openpyxl as op

def display():
    wb = op.load_workbook("ordersDB.xlsx")
    sheet = wb.active

    for row in table.get_children():
        table.delete(row)
    for row in sheet.iter_rows(min_row=2,values_only=True):
        table.insert("", tk.END, values=row)

def input_validation():
    name = cname_entry.get()
    pro = product_entry.get()
    qua = qty_entry.get()
    pri = price_entry.get()
    
    if not name or not pro or not qua or not pri:
        messagebox.showerror("Error", "Fill all")
        return False
    if not qua.isdigit() or not pri.isdigit():
        messagebox.showerror("Error", "Must be a number only")
        return False
    return True

def save():
    if not input_validation():
        return
    
    name = cname_entry.get()
    pro = product_entry.get()
    qua = int(qty_entry.get())
    pri = int(price_entry.get())
    total = qua * pri
    
    wb = op.load_workbook("ordersDB.xlsx")
    sheet = wb.active
    new_id = sheet.max_row

    sheet.append([new_id,name,pro,qua,pri,total])
    wb.save("ordersDB.xlsx")
    messagebox.showinfo("Success", "Saved")
    display()

def select(event):
    selection = table.focus()
    values = table.item(selection,"values")

    if values:
        cname_entry.delete(0, tk.END)
        product_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        product_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])

def update():
    selection = table.focus()
    if not selection:
        messagebox.showerror("Error", "Select Please")
        return
    values = table.item(selection,"values")
    recordID = values[0]

    name = cname_entry.get()
    pro = product_entry.get()
    qua = int(qty_entry.get())
    pri = int(price_entry.get())
    total = qua * pri


    wb = op.load_workbook("ordersDB.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):
        if int(row[0].value) == int(recordID):
            row[1].value = name
            row[2].value = pro
            row[3].value = qua
            row[4].value = pri
            row[5].value = total
            
    wb.save("ordersDB.xlsx")
    display()
    messagebox.showinfo("Success", "Updated")

        
        
window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")
# Form Title
title = tk.Label(window, text="Simple Ordering System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Entry
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue")
product_label.grid(row=3, column=3, columnspan=2)

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink",command=save)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen", command=update)
update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"))
delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", select)
window.mainloop()